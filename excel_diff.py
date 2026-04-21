#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel差分比較ツール
使用方法: python excel_diff.py 正解ファイル.xlsx 自動作成ファイル.xlsx
"""

import pandas as pd
import sys
from openpyxl import load_workbook

def compare(file1, file2):
    print(f"\n{'='*60}")
    print(f"【比較】")
    print(f"  正解:     {file1}")
    print(f"  自動作成: {file2}")
    print(f"{'='*60}")

    # openpyxlで読み込み
    wb1 = load_workbook(file1, data_only=True)
    wb2 = load_workbook(file2, data_only=True)

    ws1 = wb1.active
    ws2 = wb2.active

    print(f"\n■ シート名")
    print(f"  正解:     {wb1.sheetnames}")
    print(f"  自動作成: {wb2.sheetnames}")

    print(f"\n■ サイズ")
    print(f"  正解:     {ws1.max_row}行 × {ws1.max_column}列")
    print(f"  自動作成: {ws2.max_row}行 × {ws2.max_column}列")

    # 行5のヘッダー比較
    print(f"\n■ ヘッダー行（行5）比較")
    h1 = [ws1.cell(5,c).value for c in range(1, max(ws1.max_column,ws2.max_column)+1)]
    h2 = [ws2.cell(5,c).value for c in range(1, max(ws1.max_column,ws2.max_column)+1)]
    for i,(a,b) in enumerate(zip(h1,h2),1):
        if a != b:
            print(f"  列{i:2d}: 正解={repr(a)}  自動作成={repr(b)}  ← ★違う")
        else:
            print(f"  列{i:2d}: {repr(a)}  ✓")

    # データ行の差分（最大50件）
    print(f"\n■ データ差分（行6以降、最初の違い50件まで）")
    diffs = 0
    empty1 = empty2 = matched = 0
    max_row = min(ws1.max_row, ws2.max_row)
    max_col = min(ws1.max_column, ws2.max_column)

    for r in range(6, max_row+1):
        for c in range(1, max_col+1):
            v1 = ws1.cell(r,c).value
            v2 = ws2.cell(r,c).value

            # NoneとNoneは一致
            if v1 is None and v2 is None:
                matched += 1
                continue

            # 数値の近似比較
            try:
                n1, n2 = float(str(v1).replace(',','')), float(str(v2).replace(',',''))
                if abs(n1-n2) < 0.0001:
                    matched += 1
                    continue
            except:
                pass

            # 文字列比較
            s1, s2 = str(v1).strip() if v1 is not None else '', str(v2).strip() if v2 is not None else ''
            if s1 == s2:
                matched += 1
                continue

            # 差分あり
            diffs += 1
            if diffs <= 50:
                col_letter = chr(64+c) if c<=26 else '??'
                print(f"  行{r:5d} 列{c:2d}[{col_letter}]: 正解={repr(v1)}  →  自動作成={repr(v2)}")

    print(f"\n■ 集計")
    print(f"  一致セル数: {matched:,}")
    print(f"  差分セル数: {diffs:,}")
    if diffs > 50:
        print(f"  ※ 差分が多いため先頭50件のみ表示")

    # 列ごとの差分集計
    print(f"\n■ 列ごとの差分件数（多い順）")
    col_diffs = {}
    for r in range(6, max_row+1):
        for c in range(1, max_col+1):
            v1 = ws1.cell(r,c).value
            v2 = ws2.cell(r,c).value
            if v1 is None and v2 is None: continue
            try:
                n1,n2=float(str(v1).replace(',','')),float(str(v2).replace(',',''))
                if abs(n1-n2)<0.0001: continue
            except: pass
            s1 = str(v1).strip() if v1 is not None else ''
            s2 = str(v2).strip() if v2 is not None else ''
            if s1 != s2:
                hdr = ws1.cell(5,c).value or f'列{c}'
                col_diffs[hdr] = col_diffs.get(hdr,0) + 1

    for col, cnt in sorted(col_diffs.items(), key=lambda x:-x[1])[:15]:
        bar = '█' * min(cnt//100+1, 30)
        print(f"  {str(col):20s}: {cnt:5,}件  {bar}")

    print(f"\n{'='*60}\n")

if __name__ == '__main__':
    if len(sys.argv) < 3:
        print("使用方法: python excel_diff.py 正解ファイル.xlsx 自動作成ファイル.xlsx")
        sys.exit(1)
    compare(sys.argv[1], sys.argv[2])
