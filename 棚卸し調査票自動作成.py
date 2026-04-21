#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
棚卸調査票 自動作成スクリプト
使用方法: python 棚卸し調査票自動作成.py
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os, sys, glob

# =============================================
# 設定
# =============================================
INPUT_DIR   = "."
OUTPUT_FILE = "棚卸調査票自動作成.xlsx"
TANKO_DATE  = ""   # 例: "20250630"　空欄なら自動検出
# =============================================

def log(msg):       print(f"  {msg}")
def step(n, msg):   print(f"\n[Step {n}] {msg}")

def read_csv(path):
    for enc in ["cp932", "shift_jis", "utf-8-sig", "utf-8"]:
        try:
            df = pd.read_csv(path, encoding=enc, dtype=str, on_bad_lines="skip")
            df.columns = [c.strip() for c in df.columns]
            return df
        except (UnicodeDecodeError, LookupError):
            continue
    df = pd.read_csv(path, encoding="cp932", dtype=str,
                     on_bad_lines="skip", encoding_errors="replace")
    df.columns = [c.strip() for c in df.columns]
    return df

def load_one(pattern):
    files = sorted(glob.glob(os.path.join(INPUT_DIR, pattern)))
    if not files:
        log(f"⚠ ファイルなし: {pattern}")
        return None
    df = read_csv(files[0])
    log(f"読込: {os.path.basename(files[0])}  ({len(df)}行)")
    return df

def nk(s):
    """文字列正規化（前後空白除去）"""
    return s.astype(str).str.strip()

def to_f(s):
    return pd.to_numeric(s, errors="coerce")

def vlookup(base, base_key, ref, ref_key, cols):
    """重複なし・文字列キーで安全VLOOKUP"""
    if ref is None: return base
    avail = [c for c in cols if c in ref.columns]
    if not avail: return base
    sub = ref[[ref_key] + avail].copy()
    sub[ref_key] = nk(sub[ref_key])
    sub = sub.drop_duplicates(subset=[ref_key], keep="first")
    sub = sub.rename(columns={ref_key: base_key})
    base[base_key] = nk(base[base_key])
    result = base.merge(sub, on=base_key, how="left")
    log(f"  → 結合: {avail}")
    return result

# =============================================
step(1, "CSVファイルを読み込む")
# =============================================
df_zaik = load_one("TxZAIK*")
df_prts  = load_one("TxPRTS*")
df_tank  = load_one("TxTANK*")
df_item  = load_one("TxITEM*")

df_genka_list = []
for f in sorted(glob.glob(os.path.join(INPUT_DIR, "GENKA_SUB*"))):
    d = read_csv(f)
    log(f"読込: {os.path.basename(f)}  ({len(d)}行)")
    df_genka_list.append(d)

if df_zaik is None:
    print("❌ TxZAIK.CSV が見つかりません。")
    sys.exit(1)

# =============================================
step(2, "TxZAIK 整理・CODE昇順ソート")
# =============================================
keep = [c for c in ["HOKAN","HOKANNAME","NAME","CODE","TANI",
                     "GENKA","GENKANAME","INPUTDATE","INPUTUSER","ZAIK"]
        if c in df_zaik.columns]
df_zaik = df_zaik[keep].copy()
df_zaik["CODE"] = nk(df_zaik["CODE"])
df_zaik = df_zaik.sort_values("CODE").reset_index(drop=True)
log(f"TxZAIK: {len(df_zaik)}行")

# =============================================
step(3, "TxTANK 棚卸日を検出（フィルターはかけず全件使用）")
# =============================================
tanko = TANKO_DATE
if df_tank is not None and "EDATE" in df_tank.columns:
    df_tank["EDATE"] = nk(df_tank["EDATE"])
    if not tanko:
        dates = df_tank.loc[df_tank["EDATE"] != "99999999", "EDATE"].unique()
        dates = sorted([d for d in dates if d.isdigit()], reverse=True)
        if dates:
            tanko = dates[0]
            log(f"棚卸日を自動検出: {tanko}")
    # EDATEフィルターはかけない（全件使用）
    # 99999999を優先し、次に棚卸日以前の最新日付を使う
    df_tank["_edate_sort"] = df_tank["EDATE"].apply(
        lambda x: "9" * 20 if x == "99999999" else x)
    df_tank = df_tank.sort_values("_edate_sort", ascending=False)
    df_tank = df_tank.drop(columns=["_edate_sort"])
    log(f"TxTANK: {len(df_tank)}行（全件、EDATE降順）")

# =============================================
step(4, "GENKA_SUB 結合・積上原価計算（積上原価÷展開使用量）")
# =============================================
if df_genka_list:
    # 重複削除して1つに結合
    df_genka = pd.concat(df_genka_list, ignore_index=True).drop_duplicates()
    log(f"GENKA_SUB: {len(df_genka)}行（重複削除後）")

    # 列名確認（アイテムコード・積上原価・展開使用量）
    code_col   = "アイテムコード"
    genka_col  = "積上原価"
    tenkai_col = "展開使用量"

    if code_col in df_genka.columns and genka_col in df_genka.columns:
        df_genka[code_col] = nk(df_genka[code_col])
        g = to_f(df_genka[genka_col])
        t = to_f(df_genka[tenkai_col]) if tenkai_col in df_genka.columns else pd.Series(np.ones(len(df_genka)))
        df_genka["積上原価_calc"] = np.where(t.notna() & (t != 0), g / t, g)
        log("積上原価 ÷ 展開使用量 を計算")
        # CODE重複は先頭1件
        df_genka = df_genka[[code_col, "積上原価_calc"]].drop_duplicates(code_col, keep="first")
        df_genka = df_genka.rename(columns={code_col: "CODE"})
    else:
        log(f"⚠ GENKA_SUB 列名不一致: {list(df_genka.columns)[:8]}")
        df_genka = pd.DataFrame(columns=["CODE","積上原価_calc"])
else:
    df_genka = pd.DataFrame(columns=["CODE","積上原価_calc"])

# =============================================
step(5, "VLOOKUPで各テーブルを結合")
# =============================================

# TxPRTS → 単位数(SIYOU)、親品名(OYANAME)、親品番(CODE→KCODEキー)
# TxPRTS: KCODE=子品番(=TxZAIKのCODE), CODE=親品番, OYANAME=親品名, SIYOU=使用量
if df_prts is not None:
    log("TxPRTS結合")
    # CODEの衝突を避けるため先にリネームしてから結合
    prts_sub = df_prts[["KCODE","SIYOU","OYANAME","CODE"]].copy()
    prts_sub = prts_sub.rename(columns={
        "KCODE":   "CODE",
        "SIYOU":   "単位数",
        "OYANAME": "代表親品目名",
        "CODE":    "親品目番号"
    })
    prts_sub["CODE"] = nk(prts_sub["CODE"])
    prts_sub = prts_sub.drop_duplicates(subset=["CODE"], keep="first")
    df_zaik["CODE"] = nk(df_zaik["CODE"])
    df_zaik = df_zaik.merge(prts_sub, on="CODE", how="left")
    log("  → 結合: [単位数, 代表親品目名, 親品目番号]")

# GENKA_SUB → 積上原価_calc
if len(df_genka) > 0:
    log("GENKA_SUB結合（積上原価）")
    df_genka["CODE"] = nk(df_genka["CODE"])
    df_zaik["CODE"]  = nk(df_zaik["CODE"])
    df_zaik = df_zaik.merge(df_genka, on="CODE", how="left")

# TxITEM → 仕入先名(VENDORNAME)、費目(HIMOKU)、固定レベル(FIXLEVEL)、コード(VENDOR)
if df_item is not None:
    log("TxITEM結合")
    df_zaik = vlookup(df_zaik, "CODE", df_item, "CODE",
                      ["VENDORNAME", "HIMOKU", "FIXLEVEL", "VENDOR"])

# TxTANK → CODEキーでJOIN（EDATEフィルター済みの最新単価を使用）
if df_tank is not None:
    log("TxTANK結合")
    df_zaik = vlookup(df_zaik, "CODE", df_tank, "CODE",
                      ["PRICE", "CURRE"])

# =============================================
step(6, "新積上原価を計算")
# =============================================
genka_base = to_f(df_zaik.get("積上原価_calc", pd.Series(dtype=float)))
price_vals = to_f(df_zaik["PRICE"]) if "PRICE" in df_zaik.columns else pd.Series(dtype=float)
himoku_vals = df_zaik["HIMOKU"].astype(str) if "HIMOKU" in df_zaik.columns else pd.Series(dtype=str)
curre_vals  = df_zaik["CURRE"].astype(str)  if "CURRE"  in df_zaik.columns else pd.Series(dtype=str)

# 新積上原価: GENKA_SUBの積上原価計算値のみ。なければNone→#N/A
# ※費目622/625や外貨での上書きはしない（正解版のルール）
df_zaik["新積上原価"] = genka_base
log(f"新積上原価設定: {genka_base.notna().sum()}件")

# =============================================
step(7, "HOKAN→CODE昇順ソート")
# =============================================
# HOKAN→CODE 文字列昇順ソート
df_zaik["_hs"] = df_zaik["HOKAN"].astype(str).str.strip()
df_zaik["_cs"] = df_zaik["CODE"].astype(str).str.strip()
df_zaik = df_zaik.sort_values(["_hs","_cs"]).reset_index(drop=True)
df_zaik = df_zaik.drop(columns=["_hs","_cs"])
log(f"最終: {len(df_zaik)}行")

# =============================================
step(8, f"Excelへ出力: {OUTPUT_FILE}")
# =============================================
wb = Workbook()
ws = wb.active
ws.title = "棚卸調査票"
ws.sheet_view.showGridLines = False

def bdr():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def sc(ws, r, c, val, bold=False, sz=9, fg="000000", bg=None,
       ah="left", av="center", border=True):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font = Font(name="Meiryo", bold=bold, size=sz, color=fg)
    cell.alignment = Alignment(horizontal=ah, vertical=av)
    if bg: cell.fill = PatternFill("solid", fgColor=bg)
    if border: cell.border = bdr()
    return cell

# 棚卸日表示
try:
    td = tanko if tanko else ""
    tanko_disp = f"{td[:4]}/{td[4:6]}/{td[6:]}" if len(td)==8 else "----/--/--"
except:
    tanko_disp = "----/--/--"

# 行2〜4: タイトル
ws.row_dimensions[2].height = 18
sc(ws,2,3, f"棚卸日　：{tanko_disp}", bold=True, sz=10, border=False)
sc(ws,2,4, "<本社>", bold=True, sz=10, border=False)
ws.row_dimensions[3].height = 18
sc(ws,3,3, "作業区コード　：", sz=10, border=False)
sc(ws,3,4, "****　　棚卸調査票　****　　＜社外＞", bold=True, sz=11, border=False)
ws.row_dimensions[4].height = 16
sc(ws,4,7, "（分子/分母）", sz=9, fg="666666", border=False)

# 行5: ヘッダー（A・B列は空欄、C列から）
# 正解版フォーマット: A=在庫場所コード(データのみ)、B=在庫場所名(データのみ)
# ヘッダー行はC〜W
HDR = [None, None,
       "品目名","品目番号","棚卸数","計量単位","単位数",
       None,
       "代表親品目名","親品目番号","新積上原価",
       "GENKA","GENKANAME","INPUTDATE","INPUTUSER","ZAIK",
       "コード","仕入先名","費目","固定レベル","購買単価","通貨",None]
ws.row_dimensions[5].height = 24
for ci, h in enumerate(HDR, 1):
    sc(ws, 5, ci, h, bold=True, bg="1F3864", fg="FFFFFF", ah="center", sz=9)

# 列名→df_zaik列のマッピング
def gv(row, col):
    """安全に値取得、NaN/nan/空は None に変換"""
    if col not in row.index: return None
    v = row[col]
    if v is None: return None
    if isinstance(v, float) and np.isnan(v): return None
    if str(v).strip() in ("", "nan", "None"): return None
    return v

ALT = ["DEEAF1", "FFFFFF"]
for ridx, (_, row) in enumerate(df_zaik.iterrows()):
    r = ridx + 6
    ws.row_dimensions[r].height = 15
    bg = ALT[ridx % 2]

    # L列: 費目622/625のときだけ購買単価の%.2f文字列、それ以外はNone
    himoku_val = str(gv(row,"HIMOKU") or "")
    price_val  = gv(row,"PRICE")
    try:
        if himoku_val in ("622","625"):
            if price_val is not None and str(price_val).strip() not in ('','nan','None'):
                price_str = f"{float(price_val):.2f}"
            else:
                price_str = 0  # V=Noneのとき0（数値）
        else:
            price_str = None
    except: price_str = None

    # VLOOKUPなし判定: TxPRTSにデータなし→#N/A、TxTANKにデータなし→#N/A
    has_prts  = gv(row,"単位数") is not None or gv(row,"代表親品目名") is not None
    # TxTANKにデータあり判定: PRICEかCURREが存在する（0も有効）
    _price_raw = row["PRICE"] if "PRICE" in row.index else None
    _curre_raw = row["CURRE"] if "CURRE" in row.index else None
    def _valid(v): return v is not None and str(v).strip() not in ('','nan','None')
    has_tank = _valid(_price_raw) or _valid(_curre_raw)
    NA = "#N/A"

    vals = [
        gv(row,"HOKAN"),                              # A: 在庫場所コード
        gv(row,"HOKANNAME"),                          # B: 在庫場所名
        gv(row,"NAME"),                               # C: 品目名
        gv(row,"CODE"),                               # D: 品目番号
        None,                                         # E: 棚卸数（現場記入）
        gv(row,"TANI"),                               # F: 計量単位
        gv(row,"単位数") if has_prts else NA,          # G: 単位数
        None,                                         # H: 空白
        gv(row,"代表親品目名") if has_prts else NA,    # I: 代表親品目名
        gv(row,"親品目番号")   if has_prts else NA,    # J: 親品目番号
        (gv(row,"新積上原価") if gv(row,"新積上原価") is not None else "#N/A"),  # K: 新積上原価
        price_str,                                    # L: 費目622/625のみ購買単価%.2f
        gv(row,"GENKA"),                              # M: 原価コード
        gv(row,"GENKANAME"),                          # N: 製造担当名
        gv(row,"INPUTDATE"),                          # O: 入力日
        gv(row,"INPUTUSER"),                          # P: 入力者
        gv(row,"ZAIK"),                               # Q: 在庫数
        gv(row,"VENDOR"),                             # R: 仕入先コード
        gv(row,"VENDORNAME"),                         # S: 仕入先名
        gv(row,"HIMOKU"),                             # T: 費目
        "0",                                          # U: 固定レベル（常に0）
        # V: TxTANKにCODEがあればPRICEをそのまま（Noneの場合はNone）、なければ#N/A
        (gv(row,"PRICE") if has_tank else NA),        # V: 購買単価
        # W: TxTANKにCODEがあればCURREをそのまま、なければ#N/A
        (gv(row,"CURRE") if has_tank else NA),        # W: 通貨
    ]

    for ci, v in enumerate(vals, 1):
        # K列(11): 数値に変換、ただし#N/Aはそのまま文字列で出力
        if ci == 11 and v is not None and v != "#N/A":
            try: v = float(v)
            except: pass
        cell = sc(ws, r, ci, v, bg=bg, sz=9)
        if ci == 11:
            cell.alignment = Alignment(horizontal="right", vertical="center")

# 列幅（A〜W）
widths = [8,20,36,14,8,8,8,3,30,14,12,8,16,14,12,10,10,30,6,8,10,6,3]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = "A6"
ws.auto_filter.ref = f"A5:{get_column_letter(23)}5"

wb.save(OUTPUT_FILE)
print(f"\n✅ 完了！  →  {os.path.abspath(OUTPUT_FILE)}")
print(f"   データ件数: {len(df_zaik):,} 件\n")
